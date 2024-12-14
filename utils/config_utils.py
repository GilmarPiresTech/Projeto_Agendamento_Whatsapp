import openpyxl

def carregar_configuracoes(caminho_excel):
    """Carrega configurações de horários do Excel."""
    config = {}
    workbook = openpyxl.load_workbook(caminho_excel)
    sheet = workbook['Configurações']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        dia, inicio, fim, status = row
        config[dia] = {'inicio': inicio, 'fim': fim, 'status': status}
    return config

def salvar_horarios_no_excel(dias, inicio, fim, folgas, caminho_excel):
    """Salva ou atualiza os horários no Excel."""
    workbook = openpyxl.load_workbook(caminho_excel)
    sheet = workbook['Configurações']
    sheet.delete_rows(2, sheet.max_row)
    for i, dia in enumerate(dias):
        status = 'Folga' if dia in folgas else 'Trabalho'
        sheet.append([dia, inicio[i] if i < len(inicio) else '', fim[i] if i < len(fim) else '', status])
    workbook.save(caminho_excel)
