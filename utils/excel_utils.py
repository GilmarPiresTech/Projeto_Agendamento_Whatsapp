import openpyxl
import pandas as pd
import os

def verificar_ou_criar_excel(caminho_excel):
    """Verifica ou cria o arquivo Excel com as abas necessárias."""
    if not os.path.exists(caminho_excel):
        workbook = openpyxl.Workbook()
        workbook.create_sheet('Agendamentos')
        workbook.create_sheet('Configurações')
        workbook.create_sheet('Folgas')
        workbook.create_sheet('TiposConsulta')
        workbook.remove(workbook['Sheet'])
        workbook.save(caminho_excel)

def carregar_agendamentos(caminho_excel):
    """Carrega os agendamentos do Excel."""
    workbook = openpyxl.load_workbook(caminho_excel)
    sheet = workbook['Agendamentos']
    return [
        {
            'ID': row[0],
            'nome': row[1],
            'email': row[2],
            'celular': row[3],
            'data': row[4],
            'horario': row[5],
            'tipo_consulta': row[6],
            'comentario': row[7]
        }
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ]

def salvar_no_excel(caminho_excel, nome, email, celular, data, horario, tipo_consulta, comentario):
    """Salva um novo agendamento no Excel."""
    workbook = openpyxl.load_workbook(caminho_excel)
    sheet = workbook['Agendamentos']
    sheet.append([nome, email, celular, data, horario, tipo_consulta, comentario])
    workbook.save(caminho_excel)

def excluir_folga(data, caminho_excel):
    """Exclui uma folga específica do Excel."""
    workbook = openpyxl.load_workbook(caminho_excel)
    sheet = workbook['Folgas']
    for row in sheet.iter_rows(min_row=2):
        if str(row[0].value) == data:
            sheet.delete_rows(row[0].row)
            break
    workbook.save(caminho_excel)
