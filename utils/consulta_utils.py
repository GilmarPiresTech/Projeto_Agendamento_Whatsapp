import openpyxl
from datetime import datetime, timedelta

def carregar_tipos_consulta(caminho_excel):
    """Carrega tipos de consulta do Excel."""
    workbook = openpyxl.load_workbook(caminho_excel)
    sheet = workbook['TiposConsulta']
    return [
        {'Tipo': row[0], 'Duração': row[1], 'Valor': row[2]}
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ]

def salvar_tipo_consulta_no_excel(tipo, duracao, valor, caminho_excel):
    """Salva um novo tipo de consulta no Excel."""
    workbook = openpyxl.load_workbook(caminho_excel)
    sheet = workbook['TiposConsulta']
    sheet.append([tipo, duracao, valor])
    workbook.save(caminho_excel)

def gerar_dias_disponiveis(config, tipos_consulta, agendamentos):
    """Gera dias e horários disponíveis."""
    dias_disponiveis = {}
    hoje = datetime.today() + timedelta(days=1)
    for i in range(7):
        dia = hoje + timedelta(days=i)
        dia_semana = dia.strftime('%A')
        configuracao = config.get(dia_semana, {})
        if configuracao.get('status') == 'Trabalho':
            inicio, fim = configuracao['inicio'], configuracao['fim']
            horarios = gerar_horarios(inicio, fim, tipos_consulta)
            dias_disponiveis[dia.strftime('%Y-%m-%d')] = horarios
    return dias_disponiveis

def verificar_dia_disponivel(data, caminho_excel):
    """Verifica se o dia está disponível."""
    # Implementar lógica adicional para verificar folgas, se necessário
    return True

def verificar_horario_trabalho(data, horario, caminho_excel):
    """Verifica se o horário está dentro do expediente."""
    # Lógica para validar se o horário está entre início e fim do expediente
    return True

def verificar_disponibilidade(data, horario, caminho_excel):
    """Verifica se o horário já está ocupado."""
    agendamentos = carregar_agendamentos(caminho_excel)
    return all(a['data'] != data or a['horario'] != horario for a in agendamentos)
