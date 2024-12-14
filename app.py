from flask import Flask, render_template, request, redirect, flash, url_for
import openpyxl
import os
from datetime import datetime, timedelta
import pandas as pd
import uuid

app = Flask(__name__)
app.secret_key = 'chave_secreta_para_mensagens_flash'

# Caminho do arquivo Excel
ARQUIVO_EXCEL = os.path.join(os.path.expanduser('~'), 'Desktop', 'agendamentos.xlsx')

# Função para verificar ou criar o arquivo Excel com as abas necessárias e títulos de colunas
def verificar_ou_criar_excel():
    if not os.path.exists(ARQUIVO_EXCEL):
        workbook = openpyxl.Workbook()

        # Criação das abas necessárias
        sheet_agendamentos = workbook.create_sheet('Agendamentos')
        
        # Criação do cabeçalho com o ID incluído desde o início
        sheet_agendamentos.append(['ID', 'Nome', 'Email', 'Celular', 'Data', 'Horário', 'Tipo de Consulta', 'Comentário'])

        sheet_configuracoes = workbook.create_sheet('Configurações')
        sheet_configuracoes.append(['Dia', 'Início', 'Fim', 'Status'])

        sheet_folgas = workbook.create_sheet('Folgas')
        sheet_folgas.append(['Data'])

        sheet_tipos_consulta = workbook.create_sheet('TiposConsulta')
        sheet_tipos_consulta.append(['Tipo', 'Duração', 'Valor'])

        # Remove a planilha padrão se presente
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])

        workbook.save(ARQUIVO_EXCEL)
    else:
        # Carrega o arquivo e verifica se os cabeçalhos estão corretos
        workbook = openpyxl.load_workbook(ARQUIVO_EXCEL)
        
        if 'Agendamentos' not in workbook.sheetnames:
            sheet_agendamentos = workbook.create_sheet('Agendamentos')
            sheet_agendamentos.append(['ID', 'Nome', 'Email', 'Celular', 'Data', 'Horário', 'Tipo de Consulta', 'Comentário'])

        if 'Configurações' not in workbook.sheetnames:
            sheet_configuracoes = workbook.create_sheet('Configurações')
            sheet_configuracoes.append(['Dia', 'Início', 'Fim', 'Status'])

        if 'Folgas' not in workbook.sheetnames:
            sheet_folgas = workbook.create_sheet('Folgas')
            sheet_folgas.append(['Data'])

        if 'TiposConsulta' not in workbook.sheetnames:
            sheet_tipos_consulta = workbook.create_sheet('TiposConsulta')
            sheet_tipos_consulta.append(['Tipo', 'Duração', 'Valor'])

        workbook.save(ARQUIVO_EXCEL)

# Função para carregar os tipos de consulta do Excel
def carregar_tipos_consulta():
    verificar_ou_criar_excel()  # Garante que o arquivo exista antes de tentar carregar
    try:
        dados = pd.read_excel(ARQUIVO_EXCEL, sheet_name='TiposConsulta')
        tipos_consulta = dados.to_dict('records')

        # Log para depuração: Verificar se a duração está sendo lida corretamente
        for tipo in tipos_consulta:
            print(f"[LOG] Tipo de Consulta: {tipo['Tipo']}, Duração: {tipo['Duração']} minutos, Valor: {tipo['Valor']}")

        return tipos_consulta
    except Exception as e:
        print(f"[ERRO] Erro ao carregar tipos de consulta: {e}")
        flash('Erro ao carregar tipos de consulta.', 'error')
        return []

# Função para salvar tipos de consulta no Excel
def salvar_tipos_consulta(tipos):
    verificar_ou_criar_excel()  # Garante que o arquivo exista antes de salvar
    try:
        df = pd.DataFrame(tipos)
        with pd.ExcelWriter(ARQUIVO_EXCEL, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # Garante que os títulos estejam corretos
            df.columns = ['Tipo', 'Duração', 'Valor']
            df.to_excel(writer, sheet_name='TiposConsulta', index=False)
    except Exception as e:
        print(f"Erro ao salvar tipos de consulta: {e}")
        flash('Erro ao salvar tipos de consulta.', 'error')

# Função para carregar agendamentos do Excel
def carregar_agendamentos():
    verificar_ou_criar_excel()  # Garante que o arquivo exista
    try:
        # Carrega o Excel
        workbook = openpyxl.load_workbook(ARQUIVO_EXCEL)
        sheet = workbook['Agendamentos']
        agendamentos = []

        # Itera sobre as linhas a partir da segunda linha (para evitar o cabeçalho)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Mapeia as colunas corretamente para os campos esperados
            agendamento = {
                'ID': row[0],  # A primeira coluna deve ser o ID
                'nome': row[1],
                'email': row[2],
                'celular': row[3],
                'data': row[4],
                'horario': row[5],
                'tipo_consulta': row[6],
                'comentario': row[7] if len(row) > 7 else '',  # Garante que o comentário existe
            }
            agendamentos.append(agendamento)
        return agendamentos
    except Exception as e:
        print(f"Erro ao carregar agendamentos: {e}")
        return []

# Função que faz o salvamento após a exclusão de um agendamento.
def salvar_agendamentos(agendamentos):
    verificar_ou_criar_excel()  # Garante que o arquivo exista
    try:
        workbook = openpyxl.load_workbook(ARQUIVO_EXCEL)
        sheet = workbook['Agendamentos']

        # Limpa todos os agendamentos existentes, mantendo o cabeçalho
        sheet.delete_rows(2, sheet.max_row)

        # Reinsere os agendamentos na planilha
        for agendamento in agendamentos:
            sheet.append([
                agendamento['ID'],
                agendamento['nome'],
                agendamento['email'],
                agendamento['celular'],
                agendamento['data'],
                agendamento['horario'],
                agendamento['tipo_consulta'],
                agendamento['comentario']
            ])

        workbook.save(ARQUIVO_EXCEL)
    except Exception as e:
        print(f"Erro ao salvar agendamentos: {e}")
        flash('Erro ao salvar agendamentos.', 'error')

# Função para salvar agendamentos no Excel
def salvar_no_excel(nome, email, celular, data, horario, tipo_consulta, comentario):
    verificar_ou_criar_excel()  # Garante que o arquivo exista antes de salvar
    try:
        workbook = openpyxl.load_workbook(ARQUIVO_EXCEL)
        sheet = workbook['Agendamentos']
        
        # Adiciona os dados na linha correta sem adicionar o cabeçalho novamente
        agendamento_id = str(uuid.uuid4())
        sheet.append([agendamento_id, nome, email, celular, data, horario, tipo_consulta, comentario])
        workbook.save(ARQUIVO_EXCEL)
    except Exception as e:
        print(f"Erro ao salvar agendamento: {e}")
        flash('Erro ao salvar agendamento.', 'error')

# Função para carregar folgas do Excel
def carregar_folgas():
    verificar_ou_criar_excel()
    try:
        folgas = []
        workbook = openpyxl.load_workbook(ARQUIVO_EXCEL)
        if 'Folgas' in workbook.sheetnames:
            sheet = workbook['Folgas']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                folgas.append(row[0])
        return folgas
    except Exception as e:
        print(f"Erro ao carregar folgas: {e}")
        flash('Erro ao carregar folgas.', 'error')
        return []

# Função para salvar folgas no Excel
def marcar_folga_no_excel(data):
    verificar_ou_criar_excel()  # Garante que o arquivo exista antes de salvar
    try:
        workbook = openpyxl.load_workbook(ARQUIVO_EXCEL)
        if 'Folgas' in workbook.sheetnames:
            sheet = workbook['Folgas']
        else:
            sheet = workbook.create_sheet('Folgas')
            sheet.append(['Data'])  # Adiciona título da coluna

        # Verifica se a folga já existe para evitar duplicação
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == data:
                flash('Esta data já está marcada como folga.', 'error')
                return

        sheet.append([data])
        workbook.save(ARQUIVO_EXCEL)
    except Exception as e:
        print(f"Erro ao salvar folga: {e}")
        flash('Erro ao salvar folga.', 'error')

# Função para carregar configurações de horários do Excel
def carregar_configuracoes():
    verificar_ou_criar_excel()
    try:
        config = {
            'Segunda-feira': {'inicio': '', 'fim': '', 'status': 'Folga'},
            'Terça-feira': {'inicio': '', 'fim': '', 'status': 'Folga'},
            'Quarta-feira': {'inicio': '', 'fim': '', 'status': 'Folga'},
            'Quinta-feira': {'inicio': '', 'fim': '', 'status': 'Folga'},
            'Sexta-feira': {'inicio': '', 'fim': '', 'status': 'Folga'},
            'Sábado': {'inicio': '', 'fim': '', 'status': 'Folga'},
            'Domingo': {'inicio': '', 'fim': '', 'status': 'Folga'}
        }

        workbook = openpyxl.load_workbook(ARQUIVO_EXCEL)
        if 'Configurações' in workbook.sheetnames:
            sheet = workbook['Configurações']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                dia, inicio, fim, status = row
                if dia in config:
                    config[dia] = {'inicio': inicio or '', 'fim': fim or '', 'status': status or 'Folga'}
        return config
    except Exception as e:
        print(f"Erro ao carregar configurações: {e}")
        flash('Erro ao carregar configurações de horários.', 'error')
        return {}

# Função para verificar se o dia está disponível para agendamento
def verificar_dia_disponivel(data):
    verificar_ou_criar_excel()  # Garante que o arquivo exista
    folgas = carregar_folgas()
    if data in folgas:
        return False

    config = carregar_configuracoes()
    dia_semana = datetime.strptime(data, '%Y-%m-%d').strftime('%A')
    dias_semana_portugues = {
        'Monday': 'Segunda-feira',
        'Tuesday': 'Terça-feira',
        'Wednesday': 'Quarta-feira',
        'Thursday': 'Quinta-feira',
        'Friday': 'Sexta-feira',
        'Saturday': 'Sábado',
        'Sunday': 'Domingo'
    }
    dia_semana_pt = dias_semana_portugues[dia_semana]

    if dia_semana_pt not in config or config[dia_semana_pt]['status'] == 'Folga':
        return False

    return True

# Função para verificar se o horário está dentro do horário de trabalho
def verificar_horario_trabalho(data, horario):
    verificar_ou_criar_excel()  # Garante que o arquivo exista
    config = carregar_configuracoes()
    dia_semana = datetime.strptime(data, '%Y-%m-%d').strftime('%A')
    dias_semana_portugues = {
        'Monday': 'Segunda-feira',
        'Tuesday': 'Terça-feira',
        'Wednesday': 'Quarta-feira',
        'Thursday': 'Quinta-feira',
        'Friday': 'Sexta-feira',
        'Saturday': 'Sábado',
        'Sunday': 'Domingo'
    }
    dia_semana_pt = dias_semana_portugues[dia_semana]

    if dia_semana_pt not in config:
        return False

    inicio = config[dia_semana_pt]['inicio']
    fim = config[dia_semana_pt]['fim']

    if not inicio or not fim:
        return False

    horario_consulta = datetime.strptime(horario, '%H:%M')
    horario_inicio = datetime.strptime(inicio, '%H:%M')
    horario_fim = datetime.strptime(fim, '%H:%M')

    return horario_inicio <= horario_consulta <= horario_fim

# Função para verificar disponibilidade do horário
def verificar_disponibilidade(data, horario):
    verificar_ou_criar_excel()  # Garante que o arquivo exista
    try:
        workbook = openpyxl.load_workbook(ARQUIVO_EXCEL)
        sheet = workbook['Agendamentos']
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[3]) == data and str(row[4]) == horario:
                return False
        return True
    except Exception as e:
        print(f"Erro ao verificar disponibilidade: {e}")
        flash('Erro ao verificar disponibilidade do horário.', 'error')
        return False


# Função para gerar horários disponíveis em intervalos de 15 minutos
def gerar_horarios_disponiveis(inicio, fim, duracao_consulta, agendamentos_existentes):
    """
    Gera uma lista de horários disponíveis dentro de um intervalo de início e fim de expediente,
    levando em consideração a duração da consulta e os agendamentos existentes.
    """
    horarios = []
    horario_atual = datetime.strptime(inicio, '%H:%M')
    horario_fim = datetime.strptime(fim, '%H:%M')

    # Último horário válido para iniciar uma consulta sem ultrapassar o fim do expediente
    ultimo_horario_valido = horario_fim - timedelta(minutes=duracao_consulta)

    # Lista para armazenar horários disponíveis minuto a minuto
    horarios_disponiveis = []

    # Verificação minuto a minuto
    while horario_atual <= ultimo_horario_valido:
        horario_fim_consulta = horario_atual + timedelta(minutes=duracao_consulta)
        conflito = False
        
        # Verifica se o horário atual conflita com algum agendamento existente
        for agendamento in agendamentos_existentes:
            horario_inicio_agendamento = datetime.strptime(agendamento['horario'], '%H:%M')
            duracao_agendamento = agendamento['duracao']
            horario_fim_agendamento = horario_inicio_agendamento + timedelta(minutes=duracao_agendamento)

            # Ajuste na verificação: considera a duração da nova consulta no encaixe
            if not (horario_fim_consulta <= horario_inicio_agendamento or horario_atual >= horario_fim_agendamento):
                conflito = True
                break

        # Adiciona o horário à lista de disponíveis se não houver conflito
        if not conflito:
            horarios_disponiveis.append(horario_atual)

        # Incremento minuto a minuto para verificar a disponibilidade corretamente
        horario_atual += timedelta(minutes=1)

    # Filtra os horários para exibição em intervalos de 15 minutos
    for horario in horarios_disponiveis:
        if horario.minute % 15 == 0:  # Exibe somente os horários em intervalos de 15 minutos
            horarios.append(horario.strftime('%H:%M'))

    return horarios

# Função para gerar os próximos dias e horários disponíveis
def gerar_dias_disponiveis(config, duracao_consulta, agendamentos_existentes):
    """
    Gera os próximos dias e horários disponíveis para agendamento com base na configuração,
    duração da consulta e agendamentos existentes.
    """
    dias_disponiveis = {}
    hoje = datetime.today() + timedelta(days=1)  # Inicia a partir do dia seguinte

    # Log para depuração: Verificar a duração da consulta que está sendo passada
    print(f"[LOG] Gerando dias disponíveis com duração da consulta: {duracao_consulta} minutos")

    for i in range(7):  # Ajuste o número de dias conforme necessário
        dia = hoje + timedelta(days=i)
        dia_semana = dia.strftime('%A')
        dias_semana_portugues = {
            'Monday': 'Segunda-feira',
            'Tuesday': 'Terça-feira',
            'Wednesday': 'Quarta-feira',
            'Thursday': 'Quinta-feira',
            'Friday': 'Sexta-feira',
            'Saturday': 'Sábado',
            'Sunday': 'Domingo'
        }
        dia_semana_pt = dias_semana_portugues[dia_semana]
        configuracao = config.get(dia_semana_pt)

        # Verificação de status e configuração dos horários
        if configuracao:
            print(f"[LOG] Configuração encontrada para {dia_semana_pt}: {configuracao}")

        # Somente adiciona dias de trabalho com horários configurados
        if configuracao and configuracao['status'] == 'Trabalho':
            inicio = configuracao['inicio']
            fim = configuracao['fim']
            print(f"[LOG] Processando {dia.strftime('%Y-%m-%d')} - Início: {inicio}, Fim: {fim}")

            if inicio and fim:
                # Obtém os agendamentos existentes para a data atual com horário e duração
                agendamentos_dia = [
                    {
                        'horario': agendamento['horario'],
                        'duracao': next(
                            (tipo['Duração'] for tipo in carregar_tipos_consulta() if tipo['Tipo'] == agendamento['tipo_consulta']),
                            15  # Duração padrão caso o tipo não seja encontrado
                        )
                    }
                    for agendamento in agendamentos_existentes if agendamento['data'] == dia.strftime('%Y-%m-%d')
                ]
                print(f"[LOG] Agendamentos existentes para {dia.strftime('%Y-%m-%d')}: {agendamentos_dia}")

                # Gera os horários disponíveis com base na configuração e agendamentos existentes
                horarios = gerar_horarios_disponiveis(inicio, fim, duracao_consulta, agendamentos_dia)

                if horarios:  # Adiciona o dia apenas se houver horários disponíveis
                    dias_disponiveis[dia.strftime('%Y-%m-%d')] = horarios
                    print(f"[LOG] Horários disponíveis para {dia.strftime('%Y-%m-%d')}: {horarios}")
                else:
                    print(f"[LOG] Sem horários disponíveis para {dia.strftime('%Y-%m-%d')}")

    # Log para depuração final
    print(f"[LOG] Dias disponíveis gerados: {dias_disponiveis}")
    
    return dias_disponiveis

# Rota principal que renderiza o formulário de agendamento
@app.route('/')
def index():
    tipos_consulta = carregar_tipos_consulta()
    config = carregar_configuracoes()
    agendamentos_existentes = carregar_agendamentos()

    # Inicialize dias_disponiveis vazio, será preenchido ao selecionar um tipo de consulta
    dias_disponiveis = {}

    # Verificar se há tipos de consulta selecionados para calcular dias disponíveis
    if tipos_consulta:
        # Utilize o primeiro tipo de consulta ou um tipo padrão para a inicialização
        tipo_selecionado = tipos_consulta[0]  # Isso pode ser adaptado conforme a lógica do sistema
        duracao_consulta = tipo_selecionado['Duração']

        # Gera os dias disponíveis para o tipo de consulta selecionado
        dias_disponiveis = gerar_dias_disponiveis(config, duracao_consulta, agendamentos_existentes)

    # Debug para garantir que os dias disponíveis estão sendo gerados
    print(f"Dias Disponíveis: {dias_disponiveis}")

    return render_template(
        'form.html', 
        tipos_consulta=tipos_consulta, 
        dias_disponiveis=dias_disponiveis
    )


# Rota para submissão do agendamento
@app.route('/submit', methods=['POST'])
def submit():
    nome = request.form['nome']
    email = request.form['email']
    celular = request.form['celular']
    tipo_consulta = request.form['tipo_consulta']
    comentario = request.form['comentario']

    # Capturar a data e o horário selecionados
    data = request.form.get('selected_date')  # Adicione um campo hidden no formulário para armazenar a data selecionada
    horario = request.form.get('selected_time')  # Adicione um campo hidden no formulário para armazenar o horário selecionado

    if not nome or not email or not celular or not data or not horario or not tipo_consulta:
        flash('Por favor, preencha todos os campos obrigatórios.', 'error')
        return redirect(url_for('index'))

    if not verificar_dia_disponivel(data):
        flash('Não é possível agendar para esta data. Selecione outro dia.', 'error')
        return redirect(url_for('index'))

    if not verificar_horario_trabalho(data, horario):
        flash('Horário fora do expediente. Selecione outro horário.', 'error')
        return redirect(url_for('index'))

    if not verificar_disponibilidade(data, horario):
        flash('O horário selecionado já está ocupado. Por favor, escolha outro horário ou dia.', 'error')
        return redirect(url_for('index'))

    salvar_no_excel(nome, email, celular, data, horario, tipo_consulta, comentario)
    flash('Agendamento realizado com sucesso!', 'success')
    return redirect(url_for('index'))

# Rota para excluir uma folga específica
@app.route('/excluir_folga', methods=['POST'])
def excluir_folga_route():
    data = request.form['data']
    excluir_folga(data)
    return redirect(url_for('marcar_folga'))

# Rota para cancelar agendamentos
@app.route('/cancelar', methods=['POST'])
def cancelar_agendamento():
    agendamento_id = request.form.get('id')
    try:
        # Carrega os agendamentos
        agendamentos = carregar_agendamentos()

        # Filtra os agendamentos para remover aquele que possui o ID correspondente
        agendamentos = [a for a in agendamentos if a['ID'] != agendamento_id]

        # Salva de volta no Excel sem o agendamento cancelado
        salvar_agendamentos(agendamentos)
        
        flash('Agendamento cancelado com sucesso.', 'success')
        return redirect(url_for('calendario'))
    except Exception as e:
        print(f"Erro ao cancelar agendamento: {e}")
        flash('Erro ao cancelar agendamento. Tente novamente.', 'error')
        return redirect(url_for('calendario'))

# Função para cancelar um agendamento específico
def cancelar_agendamento(data, horario):
    verificar_ou_criar_excel()  # Garante que o arquivo exista
    try:
        workbook = openpyxl.load_workbook(ARQUIVO_EXCEL)
        sheet = workbook['Agendamentos']
        for row in sheet.iter_rows(min_row=2):
            if str(row[3].value) == data and str(row[4].value) == horario:
                sheet.delete_rows(row[0].row, 1)
                break
        workbook.save(ARQUIVO_EXCEL)
    except Exception as e:
        print(f"Erro ao cancelar agendamento: {e}")
        flash('Erro ao cancelar agendamento.', 'error')

# Função para salvar as configurações de horários no Excel
def salvar_horarios_no_excel(dias, inicio, fim, folgas):
    verificar_ou_criar_excel()  # Garante que o arquivo exista
    try:
        workbook = openpyxl.load_workbook(ARQUIVO_EXCEL)
        if 'Configurações' in workbook.sheetnames:
            sheet = workbook['Configurações']
            sheet.delete_rows(2, sheet.max_row)
        else:
            sheet = workbook.create_sheet('Configurações')
            sheet.append(['Dia', 'Início', 'Fim', 'Status'])

        for i, dia in enumerate(dias):
            status = 'Folga' if dia in folgas else 'Trabalho'
            sheet.append([dia, inicio[i] if i < len(inicio) else '', fim[i] if i < len(fim) else '', status])

        workbook.save(ARQUIVO_EXCEL)
    except Exception as e:
        print(f"Erro ao salvar configurações de horários: {e}")
        flash('Erro ao salvar configurações de horários.', 'error')

# Rota para configurar horários de trabalho
@app.route('/config-horarios', methods=['GET', 'POST'])
def config_horarios():
    if request.method == 'POST':
        dias = request.form.getlist('dia')
        horarios_inicio = request.form.getlist('inicio')
        horarios_fim = request.form.getlist('fim')
        folgas = request.form.getlist('folga')
        salvar_horarios_no_excel(dias, horarios_inicio, horarios_fim, folgas)
        flash('Configurações de horários atualizadas com sucesso!', 'success')
        return redirect(url_for('config_horarios'))
    else:
        config = carregar_configuracoes()
        return render_template('config_horarios.html', config=config)

# Rota para marcar um dia como folga
@app.route('/marcar-folga', methods=['GET', 'POST'])
def marcar_folga():
    if request.method == 'POST':
        data = request.form['data_folga']
        if not data:
            flash('Por favor, selecione uma data.', 'error')
            return redirect(url_for('marcar_folga'))
        marcar_folga_no_excel(data)
        flash(f'O dia {data} foi marcado como folga!', 'success')
        return redirect(url_for('marcar_folga'))
    else:
        folgas = carregar_folgas()
        return render_template('marcar_folga.html', folgas=folgas)   

# Função para excluir uma folga do Excel
def excluir_folga(data):
    verificar_ou_criar_excel()  # Garante que o arquivo exista
    try:
        workbook = openpyxl.load_workbook(ARQUIVO_EXCEL)
        if 'Folgas' in workbook.sheetnames:
            sheet = workbook['Folgas']
            for row in sheet.iter_rows(min_row=2):
                if str(row[0].value) == data:
                    sheet.delete_rows(row[0].row, 1)
                    workbook.save(ARQUIVO_EXCEL)
                    flash('Folga excluída com sucesso!', 'success')
                    return
        flash('Folga não encontrada.', 'error')
    except Exception as e:
        print(f"Erro ao excluir folga: {e}")
        flash('Erro ao excluir folga.', 'error')

# Rota para cadastrar tipos de consulta
@app.route('/tipos-consulta', methods=['GET', 'POST'])
def tipos_consulta():
    if request.method == 'POST':
        tipo = request.form['tipo']
        duracao = request.form['duracao']
        valor = request.form['valor']
        if not tipo or not duracao or not valor:
            flash('Por favor, preencha todos os campos.', 'error')
            return redirect(url_for('tipos_consulta'))
        salvar_tipo_consulta_no_excel(tipo, duracao, valor)
        flash('Tipo de consulta cadastrado com sucesso!', 'success')
        return redirect(url_for('tipos_consulta'))
    else:
        tipos = carregar_tipos_consulta()
        return render_template('tipos_consulta.html', tipos=tipos)

# Função para salvar tipos de consulta no Excel
def salvar_tipo_consulta_no_excel(tipo, duracao, valor):
    verificar_ou_criar_excel()  # Garante que o arquivo exista
    try:
        workbook = openpyxl.load_workbook(ARQUIVO_EXCEL)
        if 'TiposConsulta' in workbook.sheetnames:
            sheet = workbook['TiposConsulta']
        else:
            sheet = workbook.create_sheet('TiposConsulta')
            sheet.append(['Tipo', 'Duração', 'Valor'])

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == tipo:
                flash('Este tipo de consulta já está cadastrado.', 'error')
                return

        sheet.append([tipo, duracao, valor])
        workbook.save(ARQUIVO_EXCEL)
    except Exception as e:
        print(f"Erro ao salvar tipo de consulta: {e}")
        flash('Erro ao salvar tipo de consulta.', 'error')

# Rota para editar tipo de consulta
@app.route('/editar_tipo_consulta/<int:id>', methods=['GET', 'POST'])
def editar_tipo_consulta(id):
    verificar_ou_criar_excel()
    workbook = openpyxl.load_workbook(ARQUIVO_EXCEL)
    sheet = workbook['TiposConsulta']

    # Carregar dados atuais do item
    tipo_atual = {}
    if id + 2 <= sheet.max_row:
        row = list(sheet.iter_rows(min_row=id+2, max_row=id+2, values_only=True))[0]
        tipo_atual = {'Tipo': row[0], 'Duração': row[1], 'Valor': row[2]}

    if request.method == 'POST':
        # Receber os novos valores do formulário
        novo_tipo = request.form['tipo']
        nova_duracao = request.form['duracao']
        novo_valor = request.form['valor']

        # Atualizar os valores no Excel
        sheet.cell(row=id+2, column=1).value = novo_tipo
        sheet.cell(row=id+2, column=2).value = int(nova_duracao)
        sheet.cell(row=id+2, column=3).value = float(novo_valor)
        workbook.save(ARQUIVO_EXCEL)
        flash('Tipo de consulta atualizado com sucesso!', 'success')
        return redirect(url_for('tipos_consulta'))

    return render_template('editar_tipo_consulta.html', tipo=tipo_atual, id=id)

# Rota para excluir tipo de consulta
@app.route('/excluir_tipo_consulta/<int:id>', methods=['POST'])
def excluir_tipo_consulta(id):
    verificar_ou_criar_excel()  # Garante que o arquivo Excel exista antes de operar
    try:
        workbook = openpyxl.load_workbook(ARQUIVO_EXCEL)
        sheet = workbook['TiposConsulta']
        
        # Remove a linha correspondente ao ID (index 0-based)
        # Certifica-se de que a linha a ser excluída existe
        if id + 2 <= sheet.max_row:  # id + 2 por causa do cabeçalho e base zero
            sheet.delete_rows(id + 2, 1)  # Deleta a linha correspondente ao ID
            workbook.save(ARQUIVO_EXCEL)
            flash('Tipo de consulta excluído com sucesso!', 'success')
        else:
            flash('Erro ao excluir: item não encontrado.', 'error')
    except Exception as e:
        print(f"Erro ao excluir tipo de consulta: {e}")
        flash('Erro ao excluir tipo de consulta.', 'error')
    return redirect(url_for('tipos_consulta'))

# Rota para exibir o calendário de agendamentos
@app.route('/calendario')
def calendario():
    agendamentos = carregar_agendamentos()
    return render_template('calendario.html', agendamentos=agendamentos)

# Iniciar a aplicação
if __name__ == '__main__':
    app.run(debug=True)
